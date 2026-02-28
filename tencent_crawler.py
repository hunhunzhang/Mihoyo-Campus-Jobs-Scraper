import json
import time
import re
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

class TencentJobScraper:
    def __init__(self):
        self.jobs = []
        self.output_file = "tencent_campus_jobs.xlsx"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Content-Type': 'application/json'
        }
        # 腾讯岗位大类映射
        self.category_map = {
            2: "技术",
            3: "产品",
            4: "设计",
            5: "市场",
            6: "职能"
        }

    def scrape(self):
        print("--- Scraping Tencent Job List via Request API ---")
        
        page_index = 1
        page_size = 50 # 每次请求50条岗位
        total_fetched = 0
        
        try:
            while True:
                print(f"Fetching page {page_index}...")
                # projectId=1 对标 query=p_1，即应届生校园招聘项目
                url = 'https://join.qq.com/api/v1/position/searchPosition'
                payload = {'pageIndex': page_index, 'pageSize': page_size, 'projectId': 1}
                
                try:
                    res = requests.post(url, json=payload, headers=self.headers)
                except Exception as e:
                    print(f"Failed to fetch list API: {e}")
                    break
                    
                if res.status_code != 200:
                    print(f"Error: status code {res.status_code}")
                    break
                    
                data = res.json().get('data', {})
                position_list = data.get('positionList', [])
                total_count = data.get('count', 0)
                
                if not position_list:
                    print("No more jobs found.")
                    break
                    
                for pos in position_list:
                    post_id = pos.get('postId')
                    title = pos.get('positionTitle', "")
                    family_id = pos.get('positionFamily')
                    category = self.category_map.get(family_id, str(family_id))
                    nature = pos.get('projectName', "校园招聘")
                    
                    # 获取详细职责和要求
                    print(f"  -> Fetching detail for: {title}")
                    detail = self.get_job_detail(post_id)
                    time.sleep(0.3) # 保护接口，防止被ban
                    
                    # Tencent的数据中，request=任职要求，desc=工作职责
                    req = clean_text(detail.get('request', ''))
                    desc = clean_text(detail.get('desc', ''))
                    
                    # 提取学历
                    education = extract_education(req + " " + desc)
                    
                    self.jobs.append({
                        "岗位名称": title,
                        "岗位类别": category,
                        "性质": nature,
                        "学历要求": education,
                        "任职要求": req,
                        "工作职责": desc,
                        "加分项": ""
                    })
                    total_fetched += 1
                    
                print(f"Fetched {total_fetched}/{total_count} jobs.")
                
                # 若已抓取数量不小于总数，或者本页返回不足page_size，跳出
                if len(position_list) < page_size or total_fetched >= total_count:
                    break
                    
                page_index += 1
        except KeyboardInterrupt:
            print("\nUser interrupted! Saving collected jobs so far...")
        finally:
            self.save()

    def get_job_detail(self, post_id):
        url = f'https://join.qq.com/api/v1/jobDetails/getJobDetailsByPostId?postId={post_id}'
        try:
            res = requests.get(url, headers=self.headers, timeout=10)
            if res.status_code == 200:
                return res.json().get('data', {})
        except Exception:
            pass
        return {}

    def save(self):
        print(f"\n--- Saving {len(self.jobs)} jobs to {self.output_file} ---")
        if not self.jobs:
            print("No jobs to save.")
            return

        df = pd.DataFrame(self.jobs, columns=["岗位名称", "岗位类别", "性质", "学历要求", "任职要求", "工作职责", "加分项"])
        df.to_excel(self.output_file, index=False, engine='openpyxl')
        self.format_excel()
        print(f"Successfully saved to {self.output_file}.")

    def format_excel(self):
        try:
            wb = load_workbook(self.output_file)
            ws = wb.active
            
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
            content_font = Font(name='微软雅黑', size=10)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 50
            ws.column_dimensions['G'].width = 20

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    else:
                        cell.font = content_font
                        cell.alignment = top_left_align
            
            wb.save(self.output_file)
        except Exception as e:
            print(f"Formatting failed: {e}")

def extract_education(text):
    if not text: return ""
    clean = text.replace('\n', ' ')
    patterns = [r"博士", r"硕士", r"研究生", r"本科"]
    intro = clean[:800] 
    for p in patterns:
        if re.search(p, intro):
            return f"{p}及以上"
    return "不限/未提及"

def clean_text(text):
    if not text: return ""
    text = re.sub(r'<[^>]+>', '', str(text))
    text = re.sub(r'\n+', '\n', text)
    return text.strip()

if __name__ == "__main__":
    scraper = TencentJobScraper()
    scraper.scrape()
