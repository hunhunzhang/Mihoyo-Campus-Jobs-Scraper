import pandas as pd
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import time
import re

class BaiduJobScraper:
    def __init__(self):
        self.jobs = []
        self.seen_ids = set()
        self.output_file = "baidu_campus_jobs.xlsx"
        self.browser = None

    # ... (start_browser is fine) ...

    def scrape(self):
        with sync_playwright() as p:
            self.start_browser(p)
            
            page_num = 1
            max_consecutive_dup_pages = 0
            
            while True:
                print(f"Scanning page {page_num}...")
                current_url = f"{self.base_url}&page={page_num}"
                
                try:
                    self.page.goto(current_url, timeout=30000)
                    self.page.wait_for_load_state("networkidle")
                except Exception as e:
                    print(f"Error loading page {page_num}: {e}")
                    break
                
                title_elements = self.page.locator("div[class*='post-title__']").all()
                
                if not title_elements:
                    print(f"No jobs found on page {page_num}. Stopping.")
                    break
                
                print(f"Found {len(title_elements)} jobs on page {page_num}")
                
                current_page_new_jobs = 0
                for title_el in title_elements:
                    try:
                        job_id = self.process_job_card(title_el)
                        if job_id and job_id not in self.seen_ids:
                            self.seen_ids.add(job_id)
                            current_page_new_jobs += 1
                    except Exception as e:
                        print(f"Error processing job: {e}")
                
                if current_page_new_jobs == 0:
                    print("All jobs on this page are duplicates. Reached end of list or loop.")
                    max_consecutive_dup_pages += 1
                    if max_consecutive_dup_pages >= 2:
                         break
                else:
                    max_consecutive_dup_pages = 0

                # Safety break
                if page_num >= 100: 
                     print("Safety limit reached.")
                     break
                     
                page_num += 1
                time.sleep(1) 
                
            self.save()
            self.close()

    def process_job_card(self, title_el):
        full_title = title_el.inner_text().strip()
        
        # Extract ID for deduplication: (J12345)
        job_id = None
        id_match = re.search(r'（(J\d+)）', full_title)
        if id_match:
            job_id = id_match.group(1)
        
        # ... rest of processing ...
        
        card = title_el.locator("xpath=../..") 
        card_text = card.inner_text()
        
        # Simplify title (remove ID? optional)
        title = full_title
        
        # ... logic ...
        
        meta_match = re.search(r'([A-Za-z0-9&,]+)\|([^\|]+)\|([^\|]+)\|([^\|]+)\|(\d+人)\|(\d{4}-\d{2}-\d{2})', card_text)
        
        category = ""
        location = ""
        job_type = ""
        recruit_type = ""
        
        if meta_match:
            # ACG | 北京市 | 日常实习项目 | 技术 | 2人 | Date
            location = meta_match.group(2)
            recruit_type = meta_match.group(3)
            category = meta_match.group(4)
        
        # Parse Description
        responsibilities = ""
        requirements = ""
        
        lines = [line.strip() for line in card_text.split('\n') if line.strip()]
        
        req_start = -1
        resp_start = -1
        
        for i, line in enumerate(lines):
            if "工作职责" in line or "职责" in line and len(line) < 15:
                resp_start = i
            elif "任职要求" in line or "要求" in line and len(line) < 15:
                req_start = i
        
        if resp_start != -1 and req_start != -1:
            if resp_start < req_start:
                responsibilities = "\n".join(lines[resp_start+1:req_start])
                requirements = "\n".join(lines[req_start+1:])
            else:
                requirements = "\n".join(lines[req_start+1:resp_start])
                responsibilities = "\n".join(lines[resp_start+1:])
        else:
            requirements = card_text
            responsibilities = "See requirements"

        education = "不限"
        if "本科" in card_text or "学士" in card_text:
            education = "本科及以上"
        if "硕士" in card_text or "研究生" in card_text:
            education = "硕士及以上"
        if "博士" in card_text:
            education = "博士"
            
        # Only add if unique (checked in caller, but we don't return dict anymore)
        # Wait, I need to append to self.jobs HERE or return the dict?
        # The previous code appended to self.jobs.
        # I should traverse logic carefully.
        
        # If I return ID, caller decides.
        # But I need to construct the job object inside loop?
        # I will check ID uniqueness BEFORE appending.
        
        # Since I'm inside process_job_card, I can access self.seen_ids but scraping logic is outside.
        # Better: process_job_card returns ID. If ID is new, I append to self.jobs inside process_job_card OR return the job object.
        
        # Let's modify: `process_job_card` adds to `self.jobs` only if unique.
        
        if job_id and job_id in self.seen_ids:
             return job_id # Skip adding
        
        self.jobs.append({
            "岗位名称": title,
            "岗位类别": category,
            "性质": recruit_type,
            "学历要求": education,
            "任职要求": requirements,
            "工作职责": responsibilities,
            "加分项": "" 
        })
        
        return job_id

    def save(self):
        print(f"--- Saving {len(self.jobs)} jobs to {self.output_file} ---")
        df = pd.DataFrame(self.jobs)
        # Reorder columns if they exist
        cols = ["岗位名称", "岗位类别", "性质", "学历要求", "任职要求", "工作职责", "加分项"]
        # Add missing cols
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        df = df[cols]
        
        df.to_excel(self.output_file, index=False, engine='openpyxl')
        self.format_excel()
        print("Done.")

    def format_excel(self):
        try:
            wb = load_workbook(self.output_file)
            ws = wb.active
            
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
            content_font = Font(name='微软雅黑', size=10)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 60
            ws.column_dimensions['F'].width = 60
            ws.column_dimensions['G'].width = 30

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    else:
                        cell.font = content_font
                        cell.alignment = top_left_align
            
            wb.save(self.output_file)
        except Exception as e:
            print(f"Formatting failed: {e}")

    def close(self):
        if self.browser:
            self.browser.close()

if __name__ == "__main__":
    scraper = BaiduJobScraper()
    scraper.scrape()
