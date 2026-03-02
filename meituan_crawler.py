import json

import time
import re
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

class MeituanJobScraper:
    def __init__(self):
        self.jobs = []
        self.output_file = "meituan_campus_jobs.xlsx"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Content-Type': 'application/json'
        }

    def scrape(self):
        print("--- Scraping Meituan Job List via Request API ---")
        
        page_index = 1
        page_size = 50 # 每次请求50条岗位
        total_fetched = 0
        
        job_types_to_scrape = [
            # 应届生 (jobType: 1, subCode: 1, 3, 7)
            {'name': '应届生', 'payload': {'jobType': [{'code': '1', 'subCode': ['1', '3', '7']}]}},
            # 实习生 (jobType: 2, subCode: 1, 3, 6)包含转正实习、北斗实习、日常实习
            {'name': '实习生', 'payload': {'jobType': [{'code': '2', 'subCode': ['1', '3', '6']}]}}
        ]

        try:
            for job_cat in job_types_to_scrape:
                print(f"\n--- Scraping Category: {job_cat['name']} ---")
                page_index = 1
                cat_fetched = 0
                
                while True:
                    print(f"Fetching page {page_index}...")
                    url = 'https://zhaopin.meituan.com/api/official/job/getJobList'
                    
                    payload = {
                        'page': {'pageNo': page_index, 'pageSize': page_size},
                        'jobShareType': '1',
                        'keywords': '',
                        'cityList': [],
                        'department': [],
                        'jfJgList': [],
                        'specialCode': []
                    }
                    payload.update(job_cat['payload'])
                    
                    try:
                        res = requests.post(url, json=payload, headers=self.headers, timeout=10)
                    except Exception as e:
                        print(f"Failed to fetch list API: {e}")
                        break
                        
                    if res.status_code != 200:
                        print(f"Error: status code {res.status_code}")
                        break
                        
                    data = res.json().get('data', {})
                    if not data:
                        print("No data in response!")
                        break
                        
                    position_list = data.get('list', [])
                    total_count = data.get('page', {}).get('totalCount', 0)
                    
                    if not position_list:
                        print("No more jobs found for this category.")
                        break
                        
                    for pos in position_list:
                        job_id = pos.get('jobUnionId')
                        title = pos.get('name', "")
                        category = pos.get('jobFamily', "")  # e.g., "技术类"
                        
                        # 城市列表解析
                        city_list = pos.get('cityList', [])
                        cities = " ".join([c.get('name', '') for c in city_list if c.get('name')])
                        
                        nature = job_cat['name']
                        
                        # 获取详细职责和要求
                        print(f"  -> Fetching detail for: {title}")
                        detail = self.get_job_detail(job_id)
                        time.sleep(0.3) # 保护接口，防止被ban
                        
                        # 美团的数据中，jobDuty=工作职责，jobRequirement=任职要求
                        desc = clean_text(detail.get('jobDuty', ''))
                        req = clean_text(detail.get('jobRequirement', ''))
                        
                        # 提取学历
                        education = extract_education(req + " " + desc)
                        
                        self.jobs.append({
                            "岗位名称": title,
                            "岗位类别": category,
                            "工作城市": cities,
                            "性质": nature,
                            "学历要求": education,
                            "任职要求": req,
                            "工作职责": desc,
                            "加分项": ""
                        })
                        cat_fetched += 1
                        total_fetched += 1
                        
                    print(f"Fetched {cat_fetched}/{total_count} jobs in {job_cat['name']}.")
                    
                    if len(position_list) < page_size or cat_fetched >= total_count:
                        break
                        
                    page_index += 1
                
        except KeyboardInterrupt:
            print("\nUser interrupted! Saving collected jobs so far...")
        finally:
            self.save()

    def get_job_detail(self, job_id):
        url = 'https://zhaopin.meituan.com/api/official/job/getJobDetail'
        try:
            res = requests.post(url, json={'jobUnionId': job_id}, headers=self.headers, timeout=10)
            if res.status_code == 200:
                data = res.json().get('data')
                return data if data else {}
        except Exception:
            pass
        return {}

    def save(self):
        print(f"\n--- Saving {len(self.jobs)} jobs to {self.output_file} ---")
        if not self.jobs:
            print("No jobs to save.")
            return

        df = pd.DataFrame(self.jobs, columns=["岗位名称", "岗位类别", "工作城市", "性质", "学历要求", "任职要求", "工作职责", "加分项"])
        df.to_excel(self.output_file, index=False, engine='openpyxl')
        self.format_excel()
        print(f"Successfully saved to {self.output_file}.")

    def format_excel(self):
        try:
            wb = load_workbook(self.output_file)
            ws = wb.active
            
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
            content_font = Font(name='微软雅黑', size=10)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 50
            ws.column_dimensions['G'].width = 50
            ws.column_dimensions['H'].width = 20

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    else:
                        cell.font = content_font
                        cell.alignment = top_left_align
            
            wb.save(self.output_file)
        except Exception as e:
            print(f"Formatting failed: {e}")

def extract_education(text):
    if not text: return ""
    clean = text.replace('\n', ' ')
    patterns = [r"博士", r"硕士", r"研究生", r"本科"]
    intro = clean[:800] 
    for p in patterns:
        if re.search(p, intro):
            return f"{p}及以上"
    return "不限/未提及"

def clean_text(text):
    if not text: return ""
    text = re.sub(r'<[^>]+>', '', str(text))
    text = re.sub(r'\n+', '\n', text)
    return text.strip()

if __name__ == "__main__":
    scraper = MeituanJobScraper()
    scraper.scrape()
