import time
import re
import pandas as pd
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

class MihoyoJobScraper:
    def __init__(self):
        self.jobs = {} # Use dict keyed by ID to avoid duplicates
        self.browser = None
        self.context = None
        self.page = None

    def start_browser(self, p):
        # Try different browser channels
        try:
            print("Launching Chrome...")
            self.browser = p.chromium.launch(channel="chrome", headless=False)
        except Exception:
            try:
                print("Chrome not found. Launching Edge...")
                self.browser = p.chromium.launch(channel="msedge", headless=False)
            except Exception:
                print("Edge not found. Launching bundled Chromium...")
                self.browser = p.chromium.launch(headless=False)
        
        self.context = self.browser.new_context()
        self.page = self.context.new_page()

    def handle_list_response(self, response):
        """Intercepts the job list API."""
        if "ats-portal/v1/job/list" in response.url and response.request.method == "POST":
            try:
                data = response.json()
                if isinstance(data, dict):
                     payload = data.get("data", {})
                     job_list = payload.get("list", [])
                     if job_list:
                         print(f"Captured {len(job_list)} jobs from list API")
                         for job in job_list:
                             jid = str(job.get("id"))
                             if jid not in self.jobs:
                                 self.jobs[jid] = job
            except Exception:
                pass

    def handle_detail_response(self, response):
        """Intercepts the job detail API."""
        if "ats-portal/v1/job/info" in response.url:
            try:
                data = response.json()
                if data.get("code") == 0 and "data" in data:
                    job_data = data["data"]
                    jid = str(job_data.get("id"))
                    
                    if jid in self.jobs:
                        # Update existing job with details
                        self.jobs[jid].update(job_data)
                        print(f"Captured details for: {self.jobs[jid].get('title', jid)}")
                    else:
                        # Should not happen ideally if list was thorough, but just in case
                        self.jobs[jid] = job_data
            except Exception:
                pass

    def scrape_list(self):
        print("--- Phase 1: Scraping Job List ---")
        self.page.on("response", self.handle_list_response)
        
        url = "https://jobs.mihoyo.com/#/campus/position"
        self.page.goto(url)
        self.page.wait_for_timeout(3000)

        page_num = 1
        while True:
            print(f"Processing list page {page_num}...")
            self.page.wait_for_timeout(2000) # Wait for page load
            
            # Scroll to trigger lazy loads
            self.page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            
            # Pagination Logic
            try:
                # 1. Try "Next" button class
                next_btn = self.page.locator("button.btn-next")
                if next_btn.count() > 0 and next_btn.is_visible():
                    if next_btn.is_disabled():
                         print("Reached last page.")
                         break
                    next_btn.click()
                    page_num += 1
                    continue
                
                # 2. Try numeric pagination
                next_page_num = str(page_num + 1)
                next_link = self.page.get_by_text(next_page_num, exact=True)
                if next_link.count() > 0 and next_link.first.is_visible():
                     next_link.first.click()
                     page_num += 1
                     continue

                print("No more pages found.")
                break
            except Exception as e:
                print(f"Pagination error: {e}")
                break
        
        # Remove list listener to avoid noise
        self.page.remove_listener("response", self.handle_list_response)

    def scrape_details(self):
        print(f"--- Phase 2: Scraping Details for {len(self.jobs)} Jobs ---")
        self.page.on("response", self.handle_detail_response)
        
        job_ids = list(self.jobs.keys())
        for i, jid in enumerate(job_ids):
            job = self.jobs[jid]
            # Skip if already has info (unlikely from list view)
            if job.get("description"):
                continue

            url = f"https://jobs.mihoyo.com/#/campus/position/{jid}"
            try:
                self.page.goto(url)
                
                # Wait for the data to populate
                # We check if 'description' key exists in the shared dict
                start_time = time.time()
                while time.time() - start_time < 5:
                    if self.jobs[jid].get("description"):
                        break
                    self.page.wait_for_timeout(200)
                
            except Exception as e:
                print(f"Error visiting {url}: {e}")

            if (i + 1) % 10 == 0:
                print(f"Progress: {i+1}/{len(job_ids)}")

        self.page.remove_listener("response", self.handle_detail_response)

    def close(self):
        if self.browser:
            self.browser.close()

def extract_education(text):
    if not text: return ""
    clean = text.replace('\n', ' ')
    patterns = [r"博士", r"硕士", r"研究生", r"本科"]
    intro = clean[:100] # Usually in first few lines
    for p in patterns:
        if re.search(p, intro):
            return f"{p}及以上"
    return "不限/未提及"

def clean_text(text):
    if not text: return ""
    return text.strip()

def save_to_excel(jobs_data, filename):
    print(f"--- Phase 3: Saving to Excel {filename} ---")
    
    # Transform to DataFrame structure
    rows = []
    for job in jobs_data:
        req = clean_text(job.get("jobRequire", ""))
        rows.append({
            "岗位名称": job.get("title", ""),
            "岗位类别": job.get("competencyType", ""),
            "性质": job.get("projectName", "") or job.get("jobNature", ""),
            "学历要求": extract_education(req),
            "任职要求": req,
            "工作职责": clean_text(job.get("description", "")),
            "加分项": clean_text(job.get("addition", ""))
        })
    
    columns = ["岗位名称", "岗位类别", "性质", "学历要求", "任职要求", "工作职责", "加分项"]
    df = pd.DataFrame(rows, columns=columns)
    
    # Save raw
    df.to_excel(filename, index=False, engine='openpyxl')
    
    # Format
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        # Styles
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
        content_font = Font(name='微软雅黑', size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Column Widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 60
        ws.column_dimensions['G'].width = 30

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                else:
                    cell.font = content_font
                    cell.alignment = top_left_align
        
        wb.save(filename)
        print("Excel saved and formatted successfully.")
    except Exception as e:
        print(f"Formatting failed: {e}")

def run_crawler():
    scraper = MihoyoJobScraper()
    output_file = "e:/pachong/mihoyo_campus_jobs_full.xlsx"
    
    with sync_playwright() as p:
        scraper.start_browser(p)
        try:
            scraper.scrape_list()
            print(f"Total jobs found: {len(scraper.jobs)}")
            if len(scraper.jobs) > 0:
                #scraper.scrape_details() # Uncomment to run full detail scrape
                # For testing, we might want to just run it:
                scraper.scrape_details()
            
            # Sort keys to ensure consistent order (optional)
            sorted_jobs = [scraper.jobs[k] for k in sorted(scraper.jobs.keys())]
            save_to_excel(sorted_jobs, output_file)
            
        finally:
            scraper.close()

if __name__ == "__main__":
    run_crawler()
