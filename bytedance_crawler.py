import json
import time
import re
import pandas as pd
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

class BytedanceJobScraper:
    def __init__(self):
        self.jobs = {} 
        self.browser = None
        self.context = None
        self.page = None
        self.output_file = "bytedance_campus_jobs.xlsx"

    def start_browser(self, p):
        try:
            print("Launching Chrome...")
            self.browser = p.chromium.launch(channel="chrome", headless=False)
        except Exception:
            try:
                print("Chrome not found. Launching Edge...")
                self.browser = p.chromium.launch(channel="msedge", headless=False)
            except Exception:
                print("Edge not found. Launching bundled Chromium...")
                self.browser = p.chromium.launch(headless=False)
        
        self.context = self.browser.new_context()
        self.page = self.context.new_page()

    def handle_response(self, response):
        """Intercepts the job list API."""
        if "search/job/posts" in response.url and response.request.method in ["POST", "GET"]:
            try:
                data = response.json()
                if isinstance(data, dict):
                     payload = data.get("data", {})
                     job_list = payload.get("job_post_list", [])
                     if job_list:
                         print(f"Captured {len(job_list)} jobs from API")
                         for job in job_list:
                             jid = str(job.get("id"))
                             if jid not in self.jobs:
                                 self.jobs[jid] = job
            except Exception:
                pass

    def simplify_job_data(self):
        """Convert raw JSON jobs to simple list for DataFrame."""
        simplified = []
        for jid, job in self.jobs.items():
            # 1. Title
            title = job.get("title", "")
            
            # 2. Category
            category = ""
            if job.get("job_category"):
                category = job["job_category"].get("name", "")
            
            # 3. Nature (Subject/Project)
            nature = ""
            subject = job.get("job_subject")
            if subject and isinstance(subject.get("name"), dict):
                nature = subject["name"].get("zh_cn", "")
            elif job.get("recruit_type"):
                nature = job["recruit_type"].get("name", "")
            
            # 4. Requirements & Description
            req = clean_text(job.get("requirement", ""))
            desc = clean_text(job.get("description", ""))
            
            # 5. Education
            education = extract_education(req)
            
            # 6. Addition (Not present in explicit field, check text?)
            # Leaving empty as per previous logic unless we parse text.
            addition = ""

            simplified.append({
                "岗位名称": title,
                "岗位类别": category,
                "性质": nature,
                "学历要求": education,
                "任职要求": req,
                "工作职责": desc,
                "加分项": addition
            })
        return simplified

    def scrape(self):
        print("--- Scraping Bytedance Job List ---")
        self.page.on("response", self.handle_response)
        
        url = "https://jobs.bytedance.com/campus/position"
        self.page.goto(url)
        self.page.wait_for_timeout(5000)

        page_num = 1
        last_job_count = 0
        no_change_count = 0
        
        while True:
            print(f"Scanning page {page_num}...")
            
            current_job_count = len(self.jobs)
            if current_job_count > last_job_count:
                no_change_count = 0
                last_job_count = current_job_count
            else:
                no_change_count += 1
            
            if no_change_count >= 5:
                # If we haven't found new jobs for 5 checks/pages, maybe we are done
                # But let's verify if we reached the end of pagination
                print("No new jobs associated with recent actions. Stopping.")
                break

            # Scroll to bottom first to trigger potential lazy loading
            self.page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            self.page.wait_for_timeout(2000)
            
            # Pagination Logic
            try:
                # Bytedance / AtsPagination specific selectors
                # Based on debug: <li title="下一页" class=" atsx-pagination-next" aria-disabled="false">
                next_btn = self.page.locator(".atsx-pagination-next, li[title='下一页']")
                
                if next_btn.count() > 0:
                    btn = next_btn.first
                    if btn.is_visible():
                        # Check disabled state
                        classes = btn.get_attribute("class") or ""
                        disabled = btn.get_attribute("aria-disabled")
                        
                        if "atsx-pagination-disabled" in classes or disabled == "true":
                            print("Next button disabled. Reached end.")
                            break
                        
                        # Click the button (or the link inside it if clicking <li> doesn't work)
                        # Usually clicking <li> works, but sometimes we need to click <a> inside
                        # Safe bet: click the element itself first
                        btn.click()
                        page_num += 1
                        
                        # Wait for API response to be captured
                        self.page.wait_for_timeout(3000) 
                        continue
                
                print("No next page button found.")
                break

            except Exception as e:
                print(f"Pagination error: {e}")
                break
        
    def save(self):
        print(f"--- Saving {len(self.jobs)} jobs to {self.output_file} ---")
        data = self.simplify_job_data()
        df = pd.DataFrame(data, columns=["岗位名称", "岗位类别", "性质", "学历要求", "任职要求", "工作职责", "加分项"])
        
        df.to_excel(self.output_file, index=False, engine='openpyxl')
        self.format_excel()
        print("Done.")

    def format_excel(self):
        try:
            wb = load_workbook(self.output_file)
            ws = wb.active
            
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
            content_font = Font(name='微软雅黑', size=10)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 60
            ws.column_dimensions['F'].width = 60
            ws.column_dimensions['G'].width = 30

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    else:
                        cell.font = content_font
                        cell.alignment = top_left_align
            
            wb.save(self.output_file)
        except Exception as e:
            print(f"Formatting failed: {e}")

    def close(self):
        if self.browser:
            self.browser.close()

def extract_education(text):
    if not text: return ""
    clean = text.replace('\n', ' ')
    patterns = [r"博士", r"硕士", r"研究生", r"本科"]
    intro = clean[:200]
    for p in patterns:
        if re.search(p, intro):
            return f"{p}及以上"
    return "不限/未提及"

def clean_text(text):
    if not text: return ""
    return text.strip()

def run_bd_crawler():
    scraper = BytedanceJobScraper()
    with sync_playwright() as p:
        scraper.start_browser(p)
        try:
            scraper.scrape()
        finally:
            scraper.close()
            scraper.save()

if __name__ == "__main__":
    run_bd_crawler()
